from fastapi import APIRouter, HTTPException, Query, Depends
from sqlalchemy.orm import Session
from sqlalchemy import or_
from database import get_db
from models_db import DBAsset, DBEmployee
from models import AssetResponse
from config import MASTER_TABLE

router = APIRouter(prefix="/api")


def _make_xlsx(sheet_title: str, headers: list, rows: list) -> bytes:
    """Build a styled XLSX workbook and return raw bytes."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name limit

    header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
    header_font = Font(name="Calibri", color="FFFFFF", bold=True, size=10)
    thin = Side(style="thin", color="D0D0D0")
    cell_border = Border(bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center")
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # Auto-fit column widths (capped at 40)
    for col_idx, header in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max(
            len(str(header)),
            max((len(str(row[col_idx - 1])) for row in rows if row), default=0),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

def _db_asset_to_response(asset: DBAsset, db: Session) -> AssetResponse:
    emp_display = ""
    if asset.assigned_to_email:
        emp = db.query(DBEmployee).filter(DBEmployee.email == asset.assigned_to_email).first()
        if emp:
            emp_display = emp.employee_display or emp.full_name or ""
            
    asset_display = f"{asset.brand or ''} {asset.model or ''} ({asset.asset_id})".strip()

    return AssetResponse(
        asset_id=asset.asset_id or "",
        asset_type=asset.asset_type or "",
        status=asset.status or "Unassigned",
        condition=asset.condition or "",
        brand=asset.brand or "",
        model=asset.model or "",
        serial_number=asset.serial_number or "",
        username=asset.assigned_to_email or "Not Assigned",
        employee_display=emp_display,
        assignment_id=asset.assignment_id or "",
        date_assigned=asset.date_assigned or "",
        location=asset.location or "",
        notes=asset.notes or "",
        storage=asset.storage or "",
        storage_2=asset.storage_2 or "",
        memory_ram=asset.memory_ram or "",
        processor=asset.processor or "",
        graphics=asset.graphics or "",
        screen_size=asset.screen_size or "",
        os=asset.os or "",
        purchase_date=asset.purchase_date or "",
        purchase_price=asset.purchase_price or "",
        vendor=asset.vendor or "",
        invoice_ref=asset.invoice_ref or "",
        warranty_end=asset.warranty_end or "",
        pin_password=asset.pin_password or "",
        charger_model=asset.charger_model or "",
        charger_serial=asset.charger_serial or "",
        charger_notes=asset.charger_notes or "",
        asset_display=asset_display,
        row_index=asset.id,
    )

@router.get("/asset/{asset_id}", response_model=AssetResponse)
async def get_asset(asset_id: str, db: Session = Depends(get_db)):
    """Lookup a single asset by AssetID or the hyphen-less QR ID (case-insensitive)."""
    upper = asset_id.strip().upper()
    clean = upper.replace("-", "")
    db_asset = (
        db.query(DBAsset)
        .filter(
            DBAsset.asset_id.ilike(upper) |
            DBAsset.asset_id.ilike(asset_id) |
            DBAsset.asset_id_qr.ilike(asset_id) |
            DBAsset.asset_id_qr.ilike(clean)
        )
        .first()
    )
    if not db_asset:
        raise HTTPException(status_code=404, detail=f"Asset '{asset_id}' not found")
    return _db_asset_to_response(db_asset, db)

@router.get("/field-values")
async def field_values(db: Session = Depends(get_db)):
    """Returns sorted unique non-empty values for brand, model, vendor, location, and asset_types."""
    brands      = db.query(DBAsset.brand).distinct().all()
    models      = db.query(DBAsset.model).distinct().all()
    vendors     = db.query(DBAsset.vendor).distinct().all()
    locations   = db.query(DBAsset.location).distinct().all()
    asset_types = db.query(DBAsset.asset_type).distinct().all()

    def clean(vals):
        res = set()
        for (v,) in vals:
            if v and str(v).lower() not in ("none", "n/a", "na", "-", ""):
                res.add(str(v).strip())
        return sorted(list(res), key=str.lower)

    return {
        "brand":       clean(brands),
        "model":       clean(models),
        "vendor":      clean(vendors),
        "location":    clean(locations),
        "asset_types": clean(asset_types),
    }

@router.get("/asset-suggestions")
async def asset_suggestions(db: Session = Depends(get_db)):
    """Distinct non-empty values for all AutoInput fields in asset forms."""
    def _vals(col):
        rows = db.query(col).distinct().all()
        result = set()
        for (v,) in rows:
            if v and str(v).strip().lower() not in ("none", "n/a", "na", "-", ""):
                result.add(str(v).strip())
        return sorted(result, key=str.lower)

    return {
        "brand":       _vals(DBAsset.brand),
        "model":       _vals(DBAsset.model),
        "vendor":      _vals(DBAsset.vendor),
        "location":    _vals(DBAsset.location),
        "processor":   _vals(DBAsset.processor),
        "storage":     _vals(DBAsset.storage),
        "storage_2":   _vals(DBAsset.storage_2),
        "memory_ram":  _vals(DBAsset.memory_ram),
        "graphics":    _vals(DBAsset.graphics),
        "screen_size": _vals(DBAsset.screen_size),
        "os":          _vals(DBAsset.os),
    }

@router.get("/assets", response_model=list[AssetResponse])
async def list_assets(
    q: str = Query(default="", description="Search across ID, brand, model, serial"),
    status: str = Query(default="", description="Filter by status (Active / In Stock)"),
    db: Session = Depends(get_db)
):
    query = db.query(DBAsset)
    
    if status:
        query = query.filter(DBAsset.status.ilike(status))
    
    if q:
        search_pattern = f"%{q}%"
        query = query.filter(
            or_(
                DBAsset.asset_id.ilike(search_pattern),
                DBAsset.brand.ilike(search_pattern),
                DBAsset.model.ilike(search_pattern),
                DBAsset.serial_number.ilike(search_pattern),
                DBAsset.asset_type.ilike(search_pattern),
                DBAsset.assigned_to_email.ilike(search_pattern)
            )
        )
        
    db_assets = query.all()
    return [_db_asset_to_response(a, db) for a in db_assets]

@router.get("/employee/{email}/assets", response_model=list[AssetResponse])
async def list_employee_assets(email: str, db: Session = Depends(get_db)):
    db_assets = db.query(DBAsset).filter(DBAsset.assigned_to_email.ilike(email.strip())).all()
    return [_db_asset_to_response(a, db) for a in db_assets]

@router.get("/assets/search", response_model=list[AssetResponse])
async def search_assets(
    employee: str = Query(default=""),
    status: str = Query(default=""),
    q: str = Query(default=""),
    db: Session = Depends(get_db)
):
    """Flexible search: filter by assigned employee, status, or text query."""
    query = db.query(DBAsset)
    if employee:
        query = query.filter(DBAsset.assigned_to_email.ilike(employee))
    if status:
        query = query.filter(DBAsset.status.ilike(status))
    if q:
        pattern = f"%{q}%"
        query = query.filter(
            or_(
                DBAsset.asset_id.ilike(pattern),
                DBAsset.brand.ilike(pattern),
                DBAsset.model.ilike(pattern),
            )
        )
    return [_db_asset_to_response(a, db) for a in query.all()]


@router.post("/asset/{asset_id}/sync")
async def sync_single_asset(asset_id: str, db: Session = Depends(get_db)):
    """Mark a single asset as pending sync (use /api/sync/push with Graph token to push)."""
    asset = db.query(DBAsset).filter(
        (DBAsset.asset_id == asset_id) | (DBAsset.asset_id_qr == asset_id)
    ).first()
    if not asset:
        raise HTTPException(status_code=404, detail=f"Asset '{asset_id}' not found")
    asset.needs_sync = True
    db.commit()
    return {"success": True, "asset_id": asset.asset_id, "message": "Marked for sync — call /api/sync/push to push to Excel."}


@router.get("/assets/export")
async def export_assets(
    q:          str = Query(default=""),
    status:     str = Query(default=""),
    asset_type: str = Query(default="", alias="type"),
    condition:  str = Query(default=""),
    employee:   str = Query(default=""),
    brand:      str = Query(default=""),
    model:      str = Query(default=""),
    from_date:  str = Query(default=""),
    to_date:    str = Query(default=""),
    fmt:        str = Query(default="csv", alias="format"),
    db: Session = Depends(get_db),
):
    """
    Export filtered asset inventory as CSV or XLSX.
    Params: q, status, type, condition, employee, brand, model, from_date, to_date, format
    """
    import csv, io
    from datetime import datetime
    from fastapi.responses import StreamingResponse, Response

    def _multi(raw: str) -> list[str]:
        return [v.strip() for v in raw.split(",") if v.strip()]

    query = db.query(DBAsset)
    if status:
        vals = _multi(status)
        query = query.filter(DBAsset.status.in_(vals) if len(vals) > 1 else DBAsset.status.ilike(vals[0]))
    if asset_type:
        vals = _multi(asset_type)
        query = query.filter(DBAsset.asset_type.in_(vals) if len(vals) > 1 else DBAsset.asset_type.ilike(vals[0]))
    if condition:
        vals = _multi(condition)
        query = query.filter(DBAsset.condition.in_(vals) if len(vals) > 1 else DBAsset.condition.ilike(vals[0]))
    if brand:
        query = query.filter(DBAsset.brand.ilike(f"%{brand}%"))
    if model:
        query = query.filter(DBAsset.model.ilike(f"%{model}%"))
    if employee:
        pattern = f"%{employee}%"
        emp_emails = [
            e.email for e in
            db.query(DBEmployee).filter(
                or_(
                    DBEmployee.email.ilike(pattern),
                    DBEmployee.full_name.ilike(pattern),
                    DBEmployee.employee_id.ilike(pattern),
                    DBEmployee.employee_display.ilike(pattern),
                )
            ).all()
        ]
        query = query.filter(
            or_(
                DBAsset.assigned_to_email.ilike(pattern),
                DBAsset.assigned_to_email.in_(emp_emails),
            )
        )
    if from_date:
        try:
            query = query.filter(DBAsset.date_assigned >= from_date)
        except Exception:
            pass
    if to_date:
        try:
            query = query.filter(DBAsset.date_assigned <= to_date)
        except Exception:
            pass
    if q:
        pattern = f"%{q}%"
        query = query.filter(
            or_(
                DBAsset.asset_id.ilike(pattern),
                DBAsset.brand.ilike(pattern),
                DBAsset.model.ilike(pattern),
                DBAsset.serial_number.ilike(pattern),
                DBAsset.asset_type.ilike(pattern),
                DBAsset.assigned_to_email.ilike(pattern),
            )
        )

    assets = query.order_by(DBAsset.asset_type, DBAsset.asset_id).all()

    # Build emp display cache
    emp_cache: dict[str, str] = {}
    for a in assets:
        if a.assigned_to_email and a.assigned_to_email not in emp_cache:
            emp = db.query(DBEmployee).filter(DBEmployee.email == a.assigned_to_email).first()
            emp_cache[a.assigned_to_email] = (emp.employee_display or emp.full_name or "") if emp else ""

    HEADERS = [
        "Asset ID", "Asset Type", "Status", "Condition", "Brand", "Model",
        "Serial Number", "Assigned To (Email)", "Assigned To (Name)", "Assignment ID",
        "Date Assigned", "Location", "Storage", "Storage 2", "RAM", "Processor",
        "Graphics", "Screen Size", "OS", "Purchase Date", "Purchase Price",
        "Vendor", "Invoice Ref", "Warranty End", "Notes",
        "Charger Model", "Charger Serial", "Charger Notes",
    ]

    def _row(a: DBAsset) -> list:
        return [
            a.asset_id or "", a.asset_type or "", a.status or "", a.condition or "",
            a.brand or "", a.model or "", a.serial_number or "",
            a.assigned_to_email or "", emp_cache.get(a.assigned_to_email or "", ""),
            a.assignment_id or "", a.date_assigned or "", a.location or "",
            a.storage or "", a.storage_2 or "", a.memory_ram or "", a.processor or "",
            a.graphics or "", a.screen_size or "", a.os or "",
            a.purchase_date or "", a.purchase_price or "",
            a.vendor or "", a.invoice_ref or "", a.warranty_end or "",
            a.notes or "", a.charger_model or "", a.charger_serial or "",
            a.charger_notes or "",
        ]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    if fmt == "xlsx":
        xlsx_bytes = _make_xlsx(f"Inventory ({ts})", HEADERS, [_row(a) for a in assets])
        return Response(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="inventory_{ts}.xlsx"'},
        )

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    for a in assets:
        writer.writerow(_row(a))
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="inventory_{ts}.csv"'},
    )
